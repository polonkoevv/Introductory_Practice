import openpyxl

import re,json
book=openpyxl.load_workbook("file.xlsx")

sheet=book.active
num_cols=sheet.max_column
num_rows=sheet.max_row
print(num_rows,num_cols)
reg=r"([А-Я])([А-Я])БО-([0-9])([0-9])-(19|20|21)"

raspisanie1={}
raspisanie2={}
raspisanie_by_day={}
subject_and_info=[]
info=[]
s=""
regf=r"(([А-Яа-яёЁ]{3,20})(-?)([А-Яа-я]+)?( +)?([А-ЯЁ][., ] ?[А-Я][., ]?)?)"


for i in range(1,num_cols):
    if re.search(reg,str(sheet.cell(row=2, column=i).value))!=None:
        group=re.search(reg,str(sheet.cell(row=2, column=i).value)).group()
        for i1 in range(4,76,2):
            if sheet.cell(row=i1, column=i).value!=None:
                s+=sheet.cell(row=i1, column=i).value
                if sheet.cell(row=i1, column=i+1).value != None:
                    s+=", " + str(sheet.cell(row=i1, column=i + 1).value)
                    if sheet.cell(row=i1, column=i + 2).value != None:
                        s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
                        if sheet.cell(row=i1, column=i + 3).value != None:
                            s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
                subject_and_info.append(s)
                s=""
            else:
                subject_and_info.append("--")
            # if sheet.cell(row=i1, column=i+2).value!=None:
            #     teacher=sheet.cell(row=i1, column=i+2).value
            #
            #     if sheet.cell(row=i1, column=i).value != None:
            #         s+=", " + str(sheet.cell(row=i1, column=i + 1).value)
            #         if sheet.cell(row=i1, column=i + 1).value != None:
            #             s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
            #             if sheet.cell(row=i1, column=i + 3).value != None:
            #                 s += ", " + str(sheet.cell(row=i1, column=i + 3).value)+str(sheet.cell(row=2, column=i).value)
            #     info.append(s)
            #     s=""
            # else:
            #     subject_and_info.append("--")

            if (i1-2)%12==0:
                raspisanie_by_day[str(sheet.cell(row=i1-10, column=1).value)]=subject_and_info

                subject_and_info=[]

        raspisanie1[str(sheet.cell(row=2, column=i).value)]=raspisanie_by_day
        raspisanie_by_day={}
        for i1 in range(5, 76, 2):
            if sheet.cell(row=i1, column=i).value != None:
                s += sheet.cell(row=i1, column=i).value
                if sheet.cell(row=i1, column=i + 1).value != None:
                    s += ", " + str(sheet.cell(row=i1, column=i + 1).value)
                    if sheet.cell(row=i1, column=i + 2).value != None:
                        s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
                        if sheet.cell(row=i1, column=i + 3).value != None:
                            s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
                subject_and_info.append(s)
                s = ""
            else:
                subject_and_info.append("--")
            if (i1 - 3) % 12 == 0:
                raspisanie_by_day[str(sheet.cell(row=i1 - 11, column=1).value)] = subject_and_info

                subject_and_info = []
        raspisanie2[str(sheet.cell(row=2, column=i).value)] = raspisanie_by_day
        raspisanie_by_day = {}
raspisanie_for_teacher1={}
raspisanie_for_teacher2={}

a=[]
for i in range(4,76,12):
    for i1 in range(6):
        a.append(" ")
    #
    # week_info[sheet.cell(row=i, column=1).value]=a
    # a=[]
def week():
    week_info = {}
    a = []
    for i in range(4, 76, 12):
        for i1 in range(6):
            a.append(" ")

        week_info[sheet.cell(row=i, column=1).value] = a
        a=[]
    return week_info


for i in range(1, num_cols):
    if sheet.cell(row=3, column=i).value != None and ("ФИО" in str(sheet.cell(row=3, column=i).value)):
        for i1 in range(4, 76):
            # if sheet.cell(row=i1, column=i).value!=None:
            # print(str(sheet.cell(row=i1, column=i).value))
            tag = re.finditer(regf, str(sheet.cell(row=i1, column=i).value))
            for name in tag:
                # print("Имя в самом начале",name.group())
                if sheet.cell(row=i1, column=i) != None:
                    # print(name.group())

                    surname = name.group().replace(",", ".")

                    surname = surname.replace(". ", '.')
                    surname = surname.replace('  ', ' ')
                    # print('surname',surname)
                    if surname[len(surname) - 1] != '.' and surname[len(surname) - 1] == '.':
                        surname += '.'
                    if surname[len(surname) - 1] == '-' or (len(surname) == 2 and surname[len(surname) - 1] == ' '):
                        continue
                    raspisanie_for_teacher1.setdefault(surname, week())
                    # print("Словарь до добавления",raspisanie_for_teacher1)
                    raspisanie_for_teacher2.setdefault(surname, week())

                    if i1 % 2 == 0:
                        # print(i1,surname)
                        # print(int(4 + 12 * int(int(i1 - 4) / int(12))),
                        #       sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value)
                        # print(((i1 - 4) % 12)/2)
                        # print("----------")
                        # if raspisanie_for_teacher1.get(surname)[
                        #
                        #     sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #     int(((i1 - 4) % 12) / 2)] \
                        #         == " " :
                        if str(sheet.cell(row=i1, column=i).value).find(surname[:3]) == 0:
                            # print("Обращение по ключу",raspisanie_for_teacher1.setdefault(surname))
                            raspisanie_for_teacher1.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 4) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  :str(sheet.cell(row=i1, column=i - 2).value).find('\n')] \
                                  + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        else:
                            raspisanie_for_teacher1.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 4) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  str(sheet.cell(row=i1, column=i - 2).value).find('\n') + 1:] + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        # print("Словарь после добавления",raspisanie_for_teacher1)
                        # print(raspisanie_for_teacher1[surname][
                        #     sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #     int(((i1 - 4) % 12) / 2)])

                        # else:
                        #     raspisanie_for_teacher1[surname][
                        #         sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #         int(((i1 - 4) % 12) / 2)] += ", " + str(
                        #         sheet.cell(row=2, column=i - 2).value)
                    else:
                        if str(sheet.cell(row=i1, column=i).value).find(surname[:3]) == 0:
                            # print("Обращение по ключу",raspisanie_for_teacher1.setdefault(surname))
                            raspisanie_for_teacher2.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 5) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  :str(sheet.cell(row=i1, column=i - 2).value).find('\n')] \
                                  + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        else:
                            raspisanie_for_teacher2.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 5) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  str(sheet.cell(row=i1, column=i - 2).value).find('\n') + 1:] + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+"\n"
book=openpyxl.load_workbook("file2.xlsx")

sheet=book.active
num_cols=sheet.max_column
num_rows=sheet.max_row


for i in range(1,num_cols):
    if re.search(reg,str(sheet.cell(row=2, column=i).value))!=None:
        group=re.search(reg,str(sheet.cell(row=2, column=i).value)).group()
        for i1 in range(4,76,2):
            if sheet.cell(row=i1, column=i).value!=None:
                s+=sheet.cell(row=i1, column=i).value
                if sheet.cell(row=i1, column=i+1).value != None:
                    s+=", " + str(sheet.cell(row=i1, column=i + 1).value)
                    if sheet.cell(row=i1, column=i + 2).value != None:
                        s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
                        if sheet.cell(row=i1, column=i + 3).value != None:
                            s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
                subject_and_info.append(s)
                s=""
            else:
                subject_and_info.append("--")


            if (i1-2)%12==0:
                raspisanie_by_day[str(sheet.cell(row=i1-10, column=1).value)]=subject_and_info

                subject_and_info=[]

        raspisanie1[str(sheet.cell(row=2, column=i).value)]=raspisanie_by_day
        raspisanie_by_day={}
        for i1 in range(5, 76, 2):
            if sheet.cell(row=i1, column=i).value != None:
                s += sheet.cell(row=i1, column=i).value
                if sheet.cell(row=i1, column=i + 1).value != None:
                    s += ", " + str(sheet.cell(row=i1, column=i + 1).value)
                    if sheet.cell(row=i1, column=i + 2).value != None:
                        s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
                        if sheet.cell(row=i1, column=i + 3).value != None:
                            s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
                subject_and_info.append(s)
                s = ""
            else:
                subject_and_info.append("--")
            if (i1 - 3) % 12 == 0:
                raspisanie_by_day[str(sheet.cell(row=i1 - 11, column=1).value)] = subject_and_info

                subject_and_info = []
        raspisanie2[str(sheet.cell(row=2, column=i).value)] = raspisanie_by_day
        raspisanie_by_day = {}

for i in range(1, num_cols):
    if sheet.cell(row=3, column=i).value != None and ("ФИО" in str(sheet.cell(row=3, column=i).value)):
        for i1 in range(4, 76):
            # if sheet.cell(row=i1, column=i).value!=None:
            # print(str(sheet.cell(row=i1, column=i).value))
            tag = re.finditer(regf, str(sheet.cell(row=i1, column=i).value))
            for name in tag:
                # print("Имя в самом начале",name.group())
                if sheet.cell(row=i1, column=i) != None:
                    # print(name.group())

                    surname = name.group().replace(",", ".")

                    surname = surname.replace(". ", '.')
                    surname = surname.replace('  ', ' ')
                    # print('surname',surname)
                    if surname[len(surname) - 1] != '.' and surname[len(surname) - 1] == '.':
                        surname += '.'
                    if surname[len(surname) - 1] == '-' or (len(surname) == 2 and surname[len(surname) - 1] == ' '):
                        continue
                    raspisanie_for_teacher1.setdefault(surname, week())
                    # print("Словарь до добавления",raspisanie_for_teacher1)
                    raspisanie_for_teacher2.setdefault(surname, week())

                    if i1 % 2 == 0:
                        # print(i1,surname)
                        # print(int(4 + 12 * int(int(i1 - 4) / int(12))),
                        #       sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value)
                        # print(((i1 - 4) % 12)/2)
                        # print("----------")
                        # if raspisanie_for_teacher1.get(surname)[
                        #
                        #     sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #     int(((i1 - 4) % 12) / 2)] \
                        #         == " " :
                        if str(sheet.cell(row=i1, column=i).value).find(surname[:3]) == 0:
                            # print("Обращение по ключу",raspisanie_for_teacher1.setdefault(surname))
                            raspisanie_for_teacher1.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 4) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  :str(sheet.cell(row=i1, column=i - 2).value).find('\n')] \
                                  + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        else:
                            raspisanie_for_teacher1.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 4) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  str(sheet.cell(row=i1, column=i - 2).value).find('\n') + 1:] + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        # print("Словарь после добавления",raspisanie_for_teacher1)
                        # print(raspisanie_for_teacher1[surname][
                        #     sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #     int(((i1 - 4) % 12) / 2)])

                        # else:
                        #     raspisanie_for_teacher1[surname][
                        #         sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #         int(((i1 - 4) % 12) / 2)] += ", " + str(
                        #         sheet.cell(row=2, column=i - 2).value)
                    else:
                        if str(sheet.cell(row=i1, column=i).value).find(surname[:3]) == 0:
                            # print("Обращение по ключу",raspisanie_for_teacher1.setdefault(surname))
                            raspisanie_for_teacher2.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 5) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  :str(sheet.cell(row=i1, column=i - 2).value).find('\n')] \
                                  + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        else:
                            raspisanie_for_teacher2.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 5) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  str(sheet.cell(row=i1, column=i - 2).value).find('\n') + 1:] + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+"\n"
book=openpyxl.load_workbook("file3.xlsx")

sheet=book.active
num_cols=sheet.max_column
num_rows=sheet.max_row


for i in range(1,num_cols):
    if re.search(reg,str(sheet.cell(row=2, column=i).value))!=None:
        group=re.search(reg,str(sheet.cell(row=2, column=i).value)).group()
        for i1 in range(4,76,2):
            if sheet.cell(row=i1, column=i).value!=None:
                s+=sheet.cell(row=i1, column=i).value
                if sheet.cell(row=i1, column=i+1).value != None:
                    s+=", " + str(sheet.cell(row=i1, column=i + 1).value)
                    if sheet.cell(row=i1, column=i + 2).value != None:
                        s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
                        if sheet.cell(row=i1, column=i + 3).value != None:
                            s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
                subject_and_info.append(s)
                s=""
            else:
                subject_and_info.append("--")
            # if sheet.cell(row=i1, column=i+2).value!=None:
            #     teacher=sheet.cell(row=i1, column=i+2).value
            #
            #     if sheet.cell(row=i1, column=i).value != None:
            #         s+=", " + str(sheet.cell(row=i1, column=i + 1).value)
            #         if sheet.cell(row=i1, column=i + 1).value != None:
            #             s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
            #             if sheet.cell(row=i1, column=i + 3).value != None:
            #                 s += ", " + str(sheet.cell(row=i1, column=i + 3).value)+str(sheet.cell(row=2, column=i).value)
            #     info.append(s)
            #     s=""
            # else:
            #     subject_and_info.append("--")

            if (i1-2)%12==0:
                raspisanie_by_day[str(sheet.cell(row=i1-10, column=1).value)]=subject_and_info

                subject_and_info=[]

        raspisanie1[str(sheet.cell(row=2, column=i).value)]=raspisanie_by_day
        raspisanie_by_day={}
        for i1 in range(5, 76, 2):
            if sheet.cell(row=i1, column=i).value != None:
                s += sheet.cell(row=i1, column=i).value
                if sheet.cell(row=i1, column=i + 1).value != None:
                    s += ", " + str(sheet.cell(row=i1, column=i + 1).value)
                    if sheet.cell(row=i1, column=i + 2).value != None:
                        s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
                        if sheet.cell(row=i1, column=i + 3).value != None:
                            s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
                subject_and_info.append(s)
                s = ""
            else:
                subject_and_info.append("--")
            if (i1 - 3) % 12 == 0:
                raspisanie_by_day[str(sheet.cell(row=i1 - 11, column=1).value)] = subject_and_info

                subject_and_info = []
        raspisanie2[str(sheet.cell(row=2, column=i).value)] = raspisanie_by_day
        raspisanie_by_day = {}


for i in range(1, num_cols):
    if sheet.cell(row=3, column=i).value != None and ("ФИО" in str(sheet.cell(row=3, column=i).value)):
        for i1 in range(4, 76):
            # if sheet.cell(row=i1, column=i).value!=None:
            # print(str(sheet.cell(row=i1, column=i).value))
            tag = re.finditer(regf, str(sheet.cell(row=i1, column=i).value))
            for name in tag:
                # print("Имя в самом начале",name.group())
                if sheet.cell(row=i1, column=i) != None:
                    # print(name.group())

                    surname = name.group().replace(",", ".")

                    surname = surname.replace(". ", '.')
                    surname = surname.replace('  ', ' ')
                    # print('surname',surname)
                    if surname[len(surname) - 1] != '.' and surname[len(surname) - 1] == '.':
                        surname += '.'
                    if surname[len(surname) - 1] == '-' or (len(surname) == 2 and surname[len(surname) - 1] == ' '):
                        continue
                    raspisanie_for_teacher1.setdefault(surname, week())
                    # print("Словарь до добавления",raspisanie_for_teacher1)
                    raspisanie_for_teacher2.setdefault(surname, week())

                    if i1 % 2 == 0:
                        # print(i1,surname)
                        # print(int(4 + 12 * int(int(i1 - 4) / int(12))),
                        #       sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value)
                        # print(((i1 - 4) % 12)/2)
                        # print("----------")
                        # if raspisanie_for_teacher1.get(surname)[
                        #
                        #     sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #     int(((i1 - 4) % 12) / 2)] \
                        #         == " " :
                        if str(sheet.cell(row=i1, column=i).value).find(surname[:3]) == 0:
                            # print("Обращение по ключу",raspisanie_for_teacher1.setdefault(surname))
                            raspisanie_for_teacher1.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 4) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  :str(sheet.cell(row=i1, column=i - 2).value).find('\n')] \
                                  + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        else:
                            raspisanie_for_teacher1.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 4) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  str(sheet.cell(row=i1, column=i - 2).value).find('\n') + 1:] + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        # print("Словарь после добавления",raspisanie_for_teacher1)
                        # print(raspisanie_for_teacher1[surname][
                        #     sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #     int(((i1 - 4) % 12) / 2)])

                        # else:
                        #     raspisanie_for_teacher1[surname][
                        #         sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                        #         int(((i1 - 4) % 12) / 2)] += ", " + str(
                        #         sheet.cell(row=2, column=i - 2).value)
                    else:
                        if str(sheet.cell(row=i1, column=i).value).find(surname[:3]) == 0:
                            # print("Обращение по ключу",raspisanie_for_teacher1.setdefault(surname))
                            raspisanie_for_teacher2.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 5) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  :str(sheet.cell(row=i1, column=i - 2).value).find('\n')] \
                                  + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+'\n'
                        else:
                            raspisanie_for_teacher2.get(surname)[
                                sheet.cell(row=int(4 + 12 * int(int(i1 - 4) / int(12))), column=1).value][
                                int(((i1 - 5) % 12) / 2)] \
                                += str(sheet.cell(row=i1, column=i - 2).value)[
                                  str(sheet.cell(row=i1, column=i - 2).value).find('\n') + 1:] + ", " + str(
                                sheet.cell(row=i1, column=i - 1).value) \
                                  + ", " + str(sheet.cell(row=i1, column=i + 1).value) + ", " + str(
                                sheet.cell(row=2, column=i - 2).value)+"\n"
# print(raspisanie_for_teacher1)
#
#
#
# print(raspisanie_for_teacher1)

# print(raspisanie1)
# print(raspisanie2)
print(raspisanie_for_teacher2)
my_file=open("raspisanie1.json","w")

my_file.write(json.dumps(raspisanie1))
my_file2=open("raspisanie2.json","w")
my_file2.write(json.dumps(raspisanie2))
my_file=open("raspisanie_for_teacher1.json","w")

my_file.write(json.dumps(raspisanie_for_teacher1))
my_file=open("raspisanie_for_teacher2.json","w")

my_file.write(json.dumps(raspisanie_for_teacher2))


























































# print(raspisanie1)
# print(raspisanie2)
# for i in raspisanie1:
#     print(i)
#     for i1 in raspisanie1[i]:
#         print(i1)
#         print(raspisanie1[i][i1])
# import openpyxl,re,json
# book=openpyxl.load_workbook("file.xlsx")
# sheet=book.active
# num_cols=sheet.max_column
# num_rows=sheet.max_row
# print(num_rows,num_cols)
# reg=r"([А-Я])([А-Я])БО-([0-9])([0-9])-21"
#
# raspisanie1={}
# raspisanie2={}
# raspisanie_by_day={}
# subject_and_info=[]
# s=""
# for i in range(1,num_cols):
#     if re.search(reg,str(sheet.cell(row=2, column=i).value))!=None:
#         group=re.search(reg,str(sheet.cell(row=2, column=i).value)).group()
#         for i1 in range(4,76,2):
#             if sheet.cell(row=i1, column=i).value!=None:
#                 s+=sheet.cell(row=i1, column=i).value
#                 if sheet.cell(row=i1, column=i+1).value != None:
#                     s+=", " + str(sheet.cell(row=i1, column=i + 1).value)
#                     if sheet.cell(row=i1, column=i + 2).value != None:
#                         s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
#                         if sheet.cell(row=i1, column=i + 3).value != None:
#                             s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
#                 subject_and_info.append(s)
#                 s=""
#             else:
#                 subject_and_info.append("--")
#
#
#             if (i1-2)%12==0:
#                 raspisanie_by_day[str(sheet.cell(row=i1-10, column=1).value)]=subject_and_info
#
#                 subject_and_info=[]
#
#         raspisanie1[str(sheet.cell(row=2, column=i).value)]=raspisanie_by_day
#         raspisanie_by_day={}
#         for i1 in range(5, 76, 2):
#             if sheet.cell(row=i1, column=i).value != None:
#                 s += sheet.cell(row=i1, column=i).value
#                 if sheet.cell(row=i1, column=i + 1).value != None:
#                     s += ", " + str(sheet.cell(row=i1, column=i + 1).value)
#                     if sheet.cell(row=i1, column=i + 2).value != None:
#                         s += ", " + str(sheet.cell(row=i1, column=i + 2).value)
#                         if sheet.cell(row=i1, column=i + 3).value != None:
#                             s += ", " + str(sheet.cell(row=i1, column=i + 3).value)
#                 subject_and_info.append(s)
#                 s = ""
#             else:
#                 subject_and_info.append("--")
#             if (i1 - 3) % 12 == 0:
#                 raspisanie_by_day[str(sheet.cell(row=i1 - 11, column=1).value)] = subject_and_info
#
#                 subject_and_info = []
#         raspisanie2[str(sheet.cell(row=2, column=i).value)] = raspisanie_by_day
#         raspisanie_by_day = {}
#
# print(raspisanie1)
# print(raspisanie2)
# #
# my_file=open("raspisanie1.json","w")
#
# my_file.write(json.dumps(raspisanie1))
# my_file2=open("raspisanie2.json","w")
# my_file2.write(json.dumps(raspisanie2))
# # print(raspisanie1)
# # print(raspisanie2)
# # for i in raspisanie1:
# #     print(i)
# #     for i1 in raspisanie1[i]:
# #         print(i1)
# #         print(raspisanie1[i][i1])




